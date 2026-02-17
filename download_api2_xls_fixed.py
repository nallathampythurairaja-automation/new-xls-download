#!/usr/bin/env python3
import os
from datetime import datetime
from typing import Any, Dict, List, Optional

import requests
import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
except Exception as e:
    openpyxl = None  # type: ignore


def _extract_rows(payload: Any, preferred_keys: List[str]) -> List[Dict[str, Any]]:
    """Extract a list-of-dicts from a CSE JSON payload."""
    if isinstance(payload, list):
        # already list of dicts
        return payload

    if isinstance(payload, dict):
        # prefer known keys
        for k in preferred_keys:
            v = payload.get(k)
            if isinstance(v, list):
                return v

        # else: first list value in the dict
        for v in payload.values():
            if isinstance(v, list):
                return v

    return []


def _prettify_excel(path: str, percent_cols: Optional[List[str]] = None) -> None:
    """Apply basic formatting so the file looks 'normal' in Excel."""
    if openpyxl is None:
        return

    percent_cols = percent_cols or []
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Header style + freeze
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Set column widths based on content length (capped)
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for c in col_cells[:2000]:  # cap scan for speed
            val = c.value
            if val is None:
                continue
            s = str(val)
            if len(s) > max_len:
                max_len = len(s)
        width = min(max(10, max_len + 2), 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Number formats: try to make key fields readable
    header = [c.value for c in ws[1]]
    for col_idx, name in enumerate(header, start=1):
        if name is None:
            continue
        name_str = str(name)

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is None:
                continue

            # percentage columns assumed to be 'percent value' like -0.0797 (meaning -0.0797%)
            if name_str in percent_cols:
                try:
                    cell.value = float(cell.value) / 100.0
                    cell.number_format = "0.00%"
                except Exception:
                    pass
            else:
                # basic numeric formatting
                if isinstance(cell.value, (int,)):
                    cell.number_format = "#,##0"
                elif isinstance(cell.value, (float,)):
                    cell.number_format = "#,##0.00"

    wb.save(path)

OUT_DIR = "data"
API_URL = "https://www.cse.lk/api/detailedTrades"

def main() -> None:
    os.makedirs(OUT_DIR, exist_ok=True)

    r = requests.post(API_URL, data={}, timeout=30)
    if r.status_code != 200:
        r = requests.get(API_URL, timeout=30)
    r.raise_for_status()

    payload = r.json()
    rows = _extract_rows(payload, preferred_keys=["detailedTrades", "reqDetailedTrades", "data"])

    if not rows:
        raise RuntimeError("CSE API returned no rows (could not find a list in JSON).")

    df = pd.DataFrame(rows)

    today = datetime.now().strftime("%Y-%m-%d")
    out_path = os.path.join(OUT_DIR, f"cse_detailed_trades_{today}.xlsx")
    df.to_excel(out_path, index=False, engine="openpyxl")

    percent_cols = [c for c in df.columns if str(c).lower() in ["changepercentage", "change_percentage", "changepercent", "change_pct"]]
    _prettify_excel(out_path, percent_cols=percent_cols)

    print("Saved:", out_path, "rows:", len(df))

if __name__ == "__main__":
    main()
